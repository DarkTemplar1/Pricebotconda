#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from pathlib import Path
import os
import unicodedata

import tkinter as tk
from tkinter import filedialog, messagebox
import tkinter.ttk as ttk

import pandas as pd
import numpy as np
import subprocess
import sys
import threading  # <-- do wątku dla Automatu / Czyszczenia

from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

APP_TITLE = "PriceBot"

# --- nazwy arkuszy ---
RAPORT_SHEET = "raport"
RAPORT_ODF = "raport_odfiltrowane"

# ---------- Helpers nazewnicze ----------

def _norm(s: str) -> str:
    return (s or "").strip().lower().replace(" ", "").replace("\xa0", "").replace("\t", "")

def _plain(s: str) -> str:
    s = (s or "").lower()
    s = "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))
    return s

def _find_col(cols, candidates):
    """Zwróć istniejącą kolumnę dopasowaną do listy kandydatów (po normalizacji / zawieraniu)."""
    norm_map = {_norm(c): c for c in cols}
    # dokładne
    for cand in candidates:
        key = _norm(cand)
        if key in norm_map:
            return norm_map[key]
    # "zawiera"
    for c in cols:
        if any(_norm(x) in _norm(c) for x in candidates):
            return c
    return None

def _trim_after_semicolon(val):
    if pd.isna(val):
        return ""
    s = str(val)
    if ";" in s:
        s = s.split(";", 1)[0].strip()
    return s

def _to_float_maybe(x):
    """Parsuje liczby typu '101,62 m²', '52 m2', '11 999 zł/m²' itd."""
    if pd.isna(x):
        return None
    s = str(x)

    # usuń jednostki
    for unit in ["m²", "m2", "zł/m²", "zł/m2", "zł"]:
        s = s.replace(unit, "")

    s = s.replace(" ", "").replace("\xa0", "")
    s = s.replace(",", ".")
    s = "".join(ch for ch in s if (ch.isdigit() or ch == "." or ch == "-"))
    try:
        return float(s) if s else None
    except Exception:
        return None

# ---------- Excel: czytaj/zapisuj TYLKO arkusz "raport" (bez kasowania innych) ----------

def _xlsx_has_sheet(path: Path, sheet_name: str) -> bool:
    try:
        wb = load_workbook(path, read_only=True, keep_vba=(path.suffix.lower() == ".xlsm"))
        return sheet_name in wb.sheetnames
    except Exception:
        return False

def _read_report_excel(path: Path, sheet_name: str = RAPORT_SHEET) -> pd.DataFrame:
    """Czyta WYŁĄCZNIE arkusz 'raport'. Jeśli nie istnieje – rzuca wyjątek."""
    if not _xlsx_has_sheet(path, sheet_name):
        raise RuntimeError(f"Plik nie zawiera arkusza '{sheet_name}'.")
    return pd.read_excel(path, sheet_name=sheet_name)

def _get_header_from_ws(ws) -> list[str]:
    header = []
    for cell in ws[1]:
        header.append(str(cell.value).strip() if cell.value is not None else "")
    while header and header[-1] == "":
        header.pop()
    return header

def ensure_raport_odfiltrowane(path: Path) -> None:
    """
    Gwarantuje istnienie arkusza 'raport_odfiltrowane' z SAMYMI nagłówkami,
    skopiowanymi z arkusza 'raport'. Nie rusza innych arkuszy.
    """
    if path.suffix.lower() not in (".xlsx", ".xlsm"):
        return

    keep_vba = path.suffix.lower() == ".xlsm"
    wb = load_workbook(path, keep_vba=keep_vba)

    # 1) zapewnij raport
    if RAPORT_SHEET not in wb.sheetnames:
        if wb.sheetnames:
            wb[wb.sheetnames[0]].title = RAPORT_SHEET
        else:
            wb.create_sheet(RAPORT_SHEET)

    ws_r = wb[RAPORT_SHEET]
    header = _get_header_from_ws(ws_r)

    # 2) zapewnij raport_odfiltrowane
    if RAPORT_ODF in wb.sheetnames:
        ws_o = wb[RAPORT_ODF]
    else:
        ws_o = wb.create_sheet(RAPORT_ODF)

    ws_o.sheet_state = "visible"

    # 3) wyczyść wszystko w arkuszu i wpisz tylko nagłówek
    if ws_o.max_row >= 1:
        ws_o.delete_rows(1, ws_o.max_row)

    # wpisz nagłówek
    for c, name in enumerate(header, start=1):
        ws_o.cell(row=1, column=c).value = name

    wb.save(path)

def _write_df_to_sheet_preserve(path: Path, df: pd.DataFrame, sheet_name: str = RAPORT_SHEET) -> None:
    """
    Zapisuje DataFrame do jednego arkusza (sheet_name) w pliku XLSX/XLSM
    i NIE dotyka pozostałych arkuszy (np. 'raport_odfiltrowane').
    Dodatkowo dopilnowuje, żeby 'raport_odfiltrowane' istniał i miał nagłówki.
    """
    keep_vba = path.suffix.lower() == ".xlsm"
    wb = load_workbook(path, keep_vba=keep_vba)

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(sheet_name)

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    wb.save(path)

    # dopilnuj raport_odfiltrowane
    try:
        ensure_raport_odfiltrowane(path)
    except Exception:
        pass

# ---------- USTAWIENIA PODGLĄDU ----------

PREVIEW_SPEC = [
    ("Nr KW",        ["Nr KW", "nr_ksiegi", "nrksiegi", "nr księgi", "numer księgi"]),
    ("Województwo",  ["Województwo", "wojewodztwo", "woj"]),
    ("Powiat",       ["Powiat"]),
    ("Gmina",        ["Gmina"]),
    ("Miejscowość",  ["Miejscowość", "Miejscowosc", "Miasto"]),
    ("Dzielnica",    ["Dzielnica", "Osiedle"]),
    ("Ulica",        ["Ulica", "Ulica(dla budynku)", "Ulica(dla lokalu)"]),
    ("Obszar",       [
        "Obszar", "metry", "powierzchnia",
        "Nr działek po średniku",
        "Nr działek", "Obręb po średniku", "Obręb"
    ]),
]

HIDDEN_PREVIEW_COLS = {_norm("Typ Księgi"), _norm("Stan Księgi")}

VALUE_COLS = [
    "Średnia cena za m2 ( z bazy)",
    "Średnia skorygowana cena za m2",
    "Statystyczna wartość nieruchomości",
]

# mapa nazw filtrów → skrypt
FILTER_SCRIPTS = {
    "Brak filtra": None,
    "Jeden właściciel": ["jeden_właściciel.py", "jeden_wlasciciel.py"],
    "LOKAL MIESZKALNY": ["LOKAL_MIESZKALNY.py", "lokal_mieszkalny.py"],
    "Jeden właściciel + LOKAL MIESZKALNY": [
        "jeden_właściciel_i_LOKAL_MIESZKALNY.py",
        "jeden_wlasciciel_i_lokal_mieszkalny.py",
    ],
    "Cofnij filtr": ["cofnij.py"],
}

# ---------- Główna klasa ----------

class App(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        self.title(APP_TITLE)
        self.minsize(980, 460)

        self.df: pd.DataFrame | None = None
        self.input_path: Path | None = None
        self.current_idx: int | None = None

        # Ustawienia
        self.input_file_var = tk.StringVar(value="")
        self.folder_var = tk.StringVar(value=str(Path.home()))      # baza: tu jest Polska.xlsx
        self.output_folder_var = tk.StringVar(value="")             # folder zapisu (Nr KW).xlsx
        self.margin_m2_var = tk.DoubleVar(value=15.0)               # okno ± m²
        self.margin_pct_var = tk.DoubleVar(value=15.0)              # obniżka % ceny
        self.filter_choice_var = tk.StringVar(value="Brak filtra")

        # --- UI ---
        root = ttk.Frame(self, padding=10)
        root.pack(fill="both", expand=True)

        # ---------- Plik wejściowy ----------
        group_in = ttk.LabelFrame(root, text="Plik raportu (wejście)")
        group_in.pack(fill="x")
        row_in = ttk.Frame(group_in)
        row_in.pack(fill="x", padx=8, pady=6)

        ttk.Entry(row_in, textvariable=self.input_file_var).pack(side="left", fill="x", expand=True)
        ttk.Button(row_in, text="Wybierz...", command=self.choose_input_file).pack(side="left", padx=(8, 0))

        # PRZYCISK CZYSZCZENIE PLIKU (kolory + logi)
        self.clean_btn = tk.Button(row_in, text="Czyszczenie Pliku", command=self.clean_input_file)
        self.clean_btn.pack(side="left", padx=(8, 0))

        # ---------- Folder bazowy ----------
        group_base = ttk.LabelFrame(root, text="Miejsce tworzenia plików i folderów")
        group_base.pack(fill="x", pady=(8, 0))
        row_base = ttk.Frame(group_base)
        row_base.pack(fill="x", padx=8, pady=6)
        ttk.Entry(row_base, textvariable=self.folder_var).pack(side="left", fill="x", expand=True)
        ttk.Button(row_base, text="Wybierz folder...", command=self.choose_base_folder).pack(side="left", padx=(8, 0))
        ttk.Button(row_base, text="Przygotowanie Aplikacji", command=self.prepare_app).pack(side="left", padx=(8, 0))

        # ---------- Baza danych ----------
        group_db = ttk.LabelFrame(root, text="Baza danych")
        group_db.pack(fill="x", pady=(8, 0))
        row_db = ttk.Frame(group_db)
        row_db.pack(fill="x", padx=8, pady=6)
        ttk.Button(row_db, text="Baza danych", command=self.run_bazadanych).pack(side="left")

        # ---------- Filtry ----------
        group_flt = ttk.LabelFrame(root, text="Filtry (opcjonalne)")
        group_flt.pack(fill="x", pady=(8, 0))
        row_flt = ttk.Frame(group_flt)
        row_flt.pack(fill="x", padx=8, pady=6)

        ttk.Label(row_flt, text="Wybierz filtr:").pack(side="left")
        cmb = ttk.Combobox(
            row_flt,
            textvariable=self.filter_choice_var,
            values=list(FILTER_SCRIPTS.keys()),
            state="readonly",
            width=35
        )
        cmb.pack(side="left", padx=(6, 6))
        cmb.current(0)
        ttk.Button(row_flt, text="Użyj filtru", command=self.apply_filter).pack(side="left")

        # ---------- Folder wyników ----------
        group_out = ttk.LabelFrame(root, text="Folder zapisu wyników")
        group_out.pack(fill="x", pady=(8, 0))
        row_out = ttk.Frame(group_out)
        row_out.pack(fill="x", padx=8, pady=6)
        ttk.Entry(row_out, textvariable=self.output_folder_var).pack(side="left", fill="x", expand=True)
        ttk.Button(row_out, text="Wybierz folder...", command=self.choose_output_folder).pack(side="left", padx=(8, 0))

        # ---------- Parametry ----------
        group_ctrl = ttk.LabelFrame(root, text="Parametry i sterowanie")
        group_ctrl.pack(fill="x", pady=(8, 0))
        row_ctrl1 = ttk.Frame(group_ctrl)
        row_ctrl1.pack(fill="x", padx=8, pady=6)

        ttk.Label(row_ctrl1, text="Pomiary brzegowe metrażu:").pack(side="left")

        ttk.Label(row_ctrl1, text="± m²:").pack(side="left", padx=(8, 2))
        ttk.Spinbox(
            row_ctrl1,
            from_=0.0, to=200.0, increment=0.5,
            width=6, textvariable=self.margin_m2_var
        ).pack(side="left")
        ttk.Label(row_ctrl1, text="obniżka % ceny:").pack(side="left", padx=(12, 2))
        ttk.Spinbox(
            row_ctrl1,
            from_=0.0, to=100.0, increment=0.5,
            width=6, textvariable=self.margin_pct_var
        ).pack(side="left")

        ttk.Button(row_ctrl1, text="‹ Poprzedni", command=self.prev_row).pack(side="left", padx=(16, 0))
        ttk.Button(row_ctrl1, text="Następny ›", command=self.next_row).pack(side="left", padx=(6, 0))
        ttk.Button(
            row_ctrl1,
            text="Oblicz i zapisz ten wiersz",
            command=self.calc_and_save_row
        ).pack(side="left", padx=(16, 0))

        self.automat_btn = tk.Button(row_ctrl1, text="Automat", command=self.automate)
        self.automat_btn.pack(side="left", padx=(6, 0))

        # ---------- Podgląd ----------
        group_preview = ttk.LabelFrame(root, text="Bieżący wiersz (podgląd)")
        group_preview.pack(fill="both", expand=True, pady=(8, 0))
        self.preview_label = ttk.Label(
            group_preview,
            text="{Wybierz plik raportu}",
            anchor="w",
            justify="left"
        )
        self.preview_label.pack(fill="both", expand=True, padx=8, pady=6)

    # ---------- uruchamianie zewnętrznych skryptów ----------

    def _run_script(self, candidates: list[str], extra_args: list[str] | None = None):
        if not candidates:
            return
        extra_args = extra_args or []
        here = Path(__file__).resolve().parent
        for name in candidates:
            script = here / name
            if script.exists():
                try:
                    subprocess.Popen(
                        [sys.executable, str(script), *extra_args],
                        cwd=str(here),
                        close_fds=(os.name != "nt"),
                        creationflags=(subprocess.CREATE_NO_WINDOW if os.name == "nt" else 0),
                    )
                    messagebox.showinfo("Uruchomiono", f"Start: {script.name}")
                    return
                except Exception as e:
                    messagebox.showerror("Błąd uruchamiania", f"Nie udało się uruchomić {script.name}:\n{e}")
                    return
        messagebox.showerror("Brak pliku", f"Nie znaleziono żadnego ze skryptów: {', '.join(candidates)}")

    # ---------- GUI actions ----------

    def choose_input_file(self):
        path = filedialog.askopenfilename(
            title="Wybierz plik raportu (CSV/XLSX/XLSM)",
            filetypes=[
                ("CSV", "*.csv"),
                ("Excel", "*.xlsx;*.xlsm"),
                ("Wszystkie pliki", "*.*"),
            ],
        )
        if not path:
            return
        self.input_file_var.set(path)
        self.input_path = Path(path)
        self.load_dataframe(self.input_path)
        self.goto_row(0)

    def load_dataframe(self, path: Path):
        try:
            if path.suffix.lower() in (".xlsx", ".xlsm"):
                # ⛔ Podgląd ma być TYLKO arkusza 'raport'
                self.df = _read_report_excel(path, sheet_name=RAPORT_SHEET)

                # ✅ dopilnuj arkusza raport_odfiltrowane (techniczne)
                try:
                    ensure_raport_odfiltrowane(path)
                except Exception:
                    pass
            else:
                # CSV nie ma arkuszy — podgląd OK
                self.df = pd.read_csv(path, sep=None, engine="python")
        except Exception as e:
            messagebox.showerror(
                "Błąd odczytu",
                f"Nie mogę wczytać arkusza '{RAPORT_SHEET}' z pliku:{path}{e}"
            )
            self.df = None
            self.current_idx = None
            self.preview_label.config(text="{Brak arkusza 'raport'}")

    # ---------- CZYSZCZENIE PLIKU Z LOGIEM ----------

    def clean_input_file(self):
        in_path = self.input_file_var.get().strip()
        if not in_path:
            messagebox.showerror("Czyszczenie", "Najpierw wybierz plik raportu (u góry).")
            return

        here = Path(__file__).resolve().parent
        candidates = ["CzyszczenieAdresu.py", "czyszczeniadresu.py"]

        script_path = None
        for name in candidates:
            p = here / name
            if p.exists():
                script_path = p
                break

        if script_path is None:
            messagebox.showerror("Czyszczenie", f"Nie znaleziono żadnego ze skryptów: {', '.join(candidates)}")
            return

        try:
            self.clean_btn.config(bg="#f7e26b", activebackground="#f5d742")
        except Exception:
            pass

        def worker():
            try:
                env = os.environ.copy()
                # Wymuś UTF-8 w procesie potomnym, żeby GUI nie wywalało się na polskich znakach
                env["PYTHONIOENCODING"] = "utf-8"

                proc = subprocess.Popen(
                    [sys.executable, str(script_path), in_path],
                    cwd=str(here),
                    stdout=subprocess.PIPE,
                    stderr=subprocess.PIPE,
                    text=True,
                    encoding="utf-8",
                    errors="replace",
                    env=env,
                    close_fds=(os.name != "nt"),
                    creationflags=(subprocess.CREATE_NO_WINDOW if os.name == "nt" else 0),
                )
                out, err = proc.communicate()
                rc = proc.returncode
            except Exception as e:
                def on_error():
                    try:
                        self.clean_btn.config(bg="#f28b82", activebackground="#ea4335")
                    except Exception:
                        pass
                    messagebox.showerror("Czyszczenie", f"Nie udało się uruchomić {script_path.name}:\n{e}")
                self.after(0, on_error)
                return

            def on_done():
                if rc == 0:
                    try:
                        self.clean_btn.config(bg="#8ef98e", activebackground="#76e476")
                    except Exception:
                        pass
                    msg = f"Zakończono działanie {script_path.name}.\nPrzetworzony plik:\n{in_path}"
                    log = (out or "").strip()
                    if log:
                        msg += "\n\nLOG:\n" + log[-800:]
                    messagebox.showinfo("Czyszczenie", msg)
                else:
                    try:
                        self.clean_btn.config(bg="#f28b82", activebackground="#ea4335")
                    except Exception:
                        pass
                    log = ((err or "") + "\n" + (out or "")).strip() or "(brak tekstu błędu na stdout/stderr)"
                    if len(log) > 1500:
                        log = "...[ucięto początek]\n" + log[-1500:]
                    messagebox.showerror("Czyszczenie", f"{script_path.name} zakończył się błędem (kod {rc}).\n\n{log}")

            self.after(0, on_done)

        threading.Thread(target=worker, daemon=True).start()


    def choose_base_folder(self):
        d = filedialog.askdirectory(title="Wybierz folder bazowy", initialdir=self.folder_var.get())
        if d:
            self.folder_var.set(d)

    def prepare_app(self):
        base = Path(self.folder_var.get()).resolve()
        for p in ["linki", "województwa", "logs"]:
            (base / p).mkdir(parents=True, exist_ok=True)

        # ✅ dopilnuj raport_odfiltrowane zawsze
        if self.input_path and self.input_path.suffix.lower() in (".xlsx", ".xlsm"):
            try:
                ensure_raport_odfiltrowane(self.input_path)
            except Exception:
                pass

        if self.input_file_var.get().strip():
            self.add_value_columns_to_input()

        messagebox.showinfo("Przygotowanie Aplikacji", f"Przygotowano strukturę w:\n{base}")

    def choose_output_folder(self):
        d = filedialog.askdirectory(
            title="Wybierz folder zapisu wyników",
            initialdir=self.output_folder_var.get() or self.folder_var.get(),
        )
        if d:
            self.output_folder_var.set(d)

    # ✅ POPRAWIONE: dodaje kolumny przez openpyxl do arkusza 'raport' bez kasowania innych arkuszy
    def add_value_columns_to_input(self):
        """
        Dodaje 3 kolumny wartości do arkusza 'raport' w pliku raportowym,
        nie kasując innych arkuszy (np. 'raport_odfiltrowane').
        Wstawia je zaraz za 'Czy udziały?' jeśli istnieje.
        Dodatkowo dopilnowuje istnienia arkusza 'raport_odfiltrowane'.
        """
        in_path_str = self.input_file_var.get().strip()
        if not in_path_str:
            messagebox.showerror("Kolumny", "Najpierw wybierz plik raportu (u góry).")
            return

        path = Path(in_path_str)
        if not path.exists():
            messagebox.showerror("Kolumny", f"Plik raportu nie istnieje:\n{path}")
            return

        if path.suffix.lower() not in (".xlsx", ".xlsm"):
            messagebox.showerror("Kolumny", "Ta funkcja działa dla plików Excel (.xlsx/.xlsm).")
            return

        try:
            keep_vba = path.suffix.lower() == ".xlsm"
            wb = load_workbook(path, keep_vba=keep_vba)

            # wybierz/utwórz arkusz 'raport'
            if RAPORT_SHEET in wb.sheetnames:
                ws = wb[RAPORT_SHEET]
            else:
                ws = wb[wb.sheetnames[0]]
                if RAPORT_SHEET in wb.sheetnames:
                    wb.remove(wb[RAPORT_SHEET])
                ws.title = RAPORT_SHEET

            # odczytaj nagłówek
            header = []
            for cell in ws[1]:
                header.append(str(cell.value).strip() if cell.value is not None else "")
            while header and header[-1] == "":
                header.pop()

            # pozycja insertu
            col_udzialy = None
            for i, name in enumerate(header):
                if _norm(name) in (_norm("Czy udziały?"), _norm("Czy udzialy")):
                    col_udzialy = i
                    break
            insert_pos = (col_udzialy + 1) if col_udzialy is not None else len(header)

            to_add = [c for c in VALUE_COLS if c not in header]
            if not to_add:
                wb.save(path)
                try:
                    ensure_raport_odfiltrowane(path)
                except Exception:
                    pass
                messagebox.showinfo("Kolumny", "Kolumny wartości już istnieją w arkuszu 'raport'.")
                return

            for k, col_name in enumerate(to_add):
                ws.insert_cols(insert_pos + 1 + k)
                ws.cell(row=1, column=insert_pos + 1 + k).value = col_name

            wb.save(path)

            # dopilnuj odfiltrowane
            try:
                ensure_raport_odfiltrowane(path)
            except Exception:
                pass

        except PermissionError:
            messagebox.showerror(
                "Kolumny",
                "Nie udało się dodać kolumn — plik jest pewnie otwarty w Excelu.\n"
                "Zamknij plik i spróbuj ponownie.",
            )
            return
        except Exception as e:
            messagebox.showerror("Kolumny", f"Nie udało się dodać kolumn:\n{e}")
            return

        messagebox.showinfo(
            "Kolumny",
            "Dodano brakujące kolumny wartości do arkusza 'raport' bez kasowania innych arkuszy.\n"
            "Dodatkowo przygotowano arkusz 'raport_odfiltrowane' (same nagłówki).",
        )

    def run_bazadanych(self):
        root_dir = Path(self.folder_var.get()).resolve()
        if not root_dir.exists():
            messagebox.showerror("Baza danych", f"Folder bazowy nie istnieje:\n{root_dir}")
            return
        try:
            from bazadanych import open_ui
        except Exception as e:
            messagebox.showerror("Baza danych", f"Nie mogę zaimportować bazadanych.py:\n{e}")
            return
        open_ui(root_dir, parent=self)

    # --------- AUTOMAT ----------

    def automate(self):
        raport = self.input_file_var.get().strip()
        baza = self.folder_var.get().strip()

        if not raport:
            messagebox.showerror("Automat", "Najpierw wybierz plik raportu (u góry).")
            return
        if not baza:
            messagebox.showerror("Automat", "Najpierw ustaw folder bazowy (tam gdzie Polska.xlsx).")
            return

        try:
            self.automat_btn.config(bg="#f7e26b", activebackground="#f5d742")
        except Exception:
            pass

        try:
            import automat
        except Exception as e:
            try:
                self.automat_btn.config(bg="", activebackground="")
            except Exception:
                pass
            messagebox.showerror("Automat", f"Nie mogę zaimportować automat.py:\n{e}")
            return

        def worker():
            try:
                rc = automat.main(["automat.py", raport, baza])
            except Exception as e:
                def on_error():
                    try:
                        self.automat_btn.config(bg="", activebackground="")
                    except Exception:
                        pass
                    messagebox.showerror("Automat", f"Błąd działania automat.py:\n{e}")
                self.after(0, on_error)
                return

            def on_done():
                if rc == 0:
                    try:
                        self.automat_btn.config(bg="#8ef98e", activebackground="#76e476")
                    except Exception:
                        pass
                    messagebox.showinfo(
                        "Automat",
                        "Zakończono działanie automat.py.\nWyniki powinny być wpisane do kolumn w raporcie.",
                    )
                else:
                    try:
                        self.automat_btn.config(bg="#f28b82", activebackground="#ea4335")
                    except Exception:
                        pass
                    messagebox.showerror("Automat", "automat.py zakończył się błędem (kod != 0). Sprawdź logi.")
            self.after(0, on_done)

        threading.Thread(target=worker, daemon=True).start()

    def apply_filter(self):
        choice = self.filter_choice_var.get()
        scripts = FILTER_SCRIPTS.get(choice)
        if not scripts:
            messagebox.showinfo("Filtry", "Wybrano 'Brak filtra' – nic nie uruchamiam.")
            return
        in_path = self.input_file_var.get().strip()
        if not in_path:
            messagebox.showerror("Filtry", "Najpierw wybierz plik raportu (u góry).")
            return
        self._run_script(scripts, extra_args=["--in", in_path])

    # ---------- Nawigacja ----------

    def prev_row(self):
        if self.df is None:
            messagebox.showinfo("Nawigacja", "Najpierw wybierz plik raportu.")
            return
        i = 0 if self.current_idx is None else max(0, self.current_idx - 1)
        self.goto_row(i)

    def next_row(self):
        if self.df is None:
            messagebox.showinfo("Nawigacja", "Najpierw wybierz plik raportu.")
            return
        n = len(self.df.index)
        i = 0 if self.current_idx is None else min(n - 1, self.current_idx + 1)
        self.goto_row(i)

    # ---------- PODGLĄD ----------

    def goto_row(self, i: int):
        if self.df is None or i < 0 or i >= len(self.df.index):
            return
        self.current_idx = i
        row = self.df.iloc[i]
        lines = [f"Wiersz {i+1}/{len(self.df)}"]
        for label, candidates in PREVIEW_SPEC:
            if any(_norm(c) in HIDDEN_PREVIEW_COLS for c in candidates):
                continue
            col = _find_col(self.df.columns, candidates)
            val = _trim_after_semicolon(row[col]) if col else ""
            lines.append(f"• {label}: {val}")
        self.preview_label.config(text="\n".join(lines))

    # ---------- KALKULACJA + ZAPIS ----------

    def calc_and_save_row(self):
        if self.df is None or self.current_idx is None:
            messagebox.showinfo("Zapis", "Najpierw wybierz plik raportu i wiersz.")
            return
        if not self.output_folder_var.get() and not self.folder_var.get():
            messagebox.showerror("Brak folderu", "Wybierz 'Folder zapisu wyników'.")
            return

        row = self.df.iloc[self.current_idx]

        # Nr KW
        kw_col = _find_col(
            self.df.columns,
            ["Nr KW", "nr_kw", "nrksiegi", "nr księgi", "nr_ksiegi", "numer księgi"],
        )
        kw_value = (
            str(row[kw_col]).strip()
            if (kw_col and pd.notna(row[kw_col]) and str(row[kw_col]).strip())
            else f"WIERSZ_{self.current_idx+1}"
        )

        # Obszar
        area_col = _find_col(self.df.columns, ["Obszar", "metry", "powierzchnia"])
        area_val = _to_float_maybe(_trim_after_semicolon(row[area_col])) if area_col else None
        if area_val is None:
            messagebox.showerror("Brak obszaru", "Nie znalazłem wartości obszaru/metry.")
            return

        def _get(cands):
            c = _find_col(self.df.columns, cands)
            return _trim_after_semicolon(row[c]) if c else ""

        woj_r = _get(["Województwo", "wojewodztwo", "woj"])
        pow_r = _get(["Powiat"])
        gmi_r = _get(["Gmina"])
        mia_r = _get(["Miejscowość", "Miejscowosc", "Miasto"])
        dzl_r = _get(["Dzielnica", "Osiedle"])
        uli_r = _get(["Ulica", "Ulica(dla budynku)", "Ulica(dla lokalu)"])

        base_dir = Path(self.folder_var.get()).resolve()
        polska_path = base_dir / "Polska.xlsx"
        if not polska_path.exists():
            messagebox.showerror("Brak pliku", f"Nie znaleziono pliku: {polska_path}")
            return
        try:
            df_pl = pd.read_excel(polska_path)
        except Exception as e:
            messagebox.showerror("Błąd odczytu", f"Nie mogę wczytać {polska_path}:\n{e}")
            return

        col_area_pl = _find_col(df_pl.columns, ["metry", "powierzchnia", "m2", "obszar"])
        col_price_pl = _find_col(df_pl.columns, ["cena_za_metr", "cena za metr", "cena za m²", "cena za m2", "cena/m2"])
        if col_area_pl is None or col_price_pl is None:
            messagebox.showerror("Kolumny w Polska.xlsx", "Nie znalazłem kolumn metrażu i/lub ceny za m² w Polska.xlsx.")
            return

        margin_m2 = float(self.margin_m2_var.get() or 0.0)
        margin_pct = float(self.margin_pct_var.get() or 0.0)

        delta = abs(margin_m2)
        low, high = max(0.0, area_val - delta), area_val + delta

        m = df_pl[col_area_pl].map(_to_float_maybe)
        mask_area = (m >= low) & (m <= high)

        def _eq_mask(col_candidates, value):
            col = _find_col(df_pl.columns, col_candidates)
            if col is None or not str(value).strip():
                return pd.Series(True, index=df_pl.index)
            s = df_pl[col].astype(str).str.strip().str.lower()
            v = str(value).strip().lower()
            return s == v

        mask_full = mask_area.copy()
        mask_full &= _eq_mask(["wojewodztwo", "województwo"], woj_r)
        mask_full &= _eq_mask(["powiat"], pow_r)
        mask_full &= _eq_mask(["gmina"], gmi_r)
        mask_full &= _eq_mask(["miejscowosc", "miasto", "miejscowość"], mia_r)
        if dzl_r:
            mask_full &= _eq_mask(["dzielnica", "osiedle"], dzl_r)
        if uli_r:
            mask_full &= _eq_mask(["ulica"], uli_r)

        df_sel = df_pl[mask_full].copy()

        if df_sel.empty and uli_r:
            mask_ul = mask_area.copy()
            mask_ul &= _eq_mask(["wojewodztwo", "województwo"], woj_r)
            mask_ul &= _eq_mask(["miejscowosc", "miasto", "miejscowość"], mia_r)
            if dzl_r:
                mask_ul &= _eq_mask(["dzielnica", "osiedle"], dzl_r)
            mask_ul &= _eq_mask(["ulica"], uli_r)
            df_sel = df_pl[mask_ul].copy()

        if df_sel.empty and dzl_r:
            mask_dziel = mask_area.copy()
            mask_dziel &= _eq_mask(["wojewodztwo", "województwo"], woj_r)
            mask_dziel &= _eq_mask(["miejscowosc", "miasto", "miejscowość"], mia_r)
            mask_dziel &= _eq_mask(["dzielnica", "osiedle"], dzl_r)
            df_sel = df_pl[mask_dziel].copy()

        if df_sel.empty and mia_r:
            mask_miasto = mask_area.copy()
            mask_miasto &= _eq_mask(["wojewodztwo", "województwo"], woj_r)
            mask_miasto &= _eq_mask(["miejscowosc", "miasto", "miejscowość"], mia_r)
            df_sel = df_pl[mask_miasto].copy()

        if df_sel.empty:
            messagebox.showinfo("Brak dopasowań", f"Nie znaleziono rekordów w zakresie [{low:.2f}; {high:.2f}] m².")
            return

        prices = df_sel[col_price_pl].map(_to_float_maybe)
        df_sel = df_sel[prices.notna()].copy()
        prices = df_sel[col_price_pl].map(_to_float_maybe)

        if len(prices) >= 4:
            q1 = np.nanpercentile(prices, 25)
            q3 = np.nanpercentile(prices, 75)
            iqr = q3 - q1
            lo = q1 - 1.5 * iqr
            hi = q3 + 1.5 * iqr
            df_sel = df_sel[(prices >= lo) & (prices <= hi)].copy()
            prices = df_sel[col_price_pl].map(_to_float_maybe)

        out_dir = Path(self.output_folder_var.get() or self.folder_var.get()).resolve()
        out_dir.mkdir(parents=True, exist_ok=True)

        safe_kw = "".join(ch for ch in kw_value if ch not in "\\/:*?\"<>|")
        out_path = out_dir / f"({safe_kw}).xlsx"

        avg = float(np.nanmean(prices)) if not df_sel.empty else None

        summary = {c: "" for c in df_sel.columns}
        summary[col_price_pl] = avg if avg is not None else ""
        df_out = pd.concat([df_sel, pd.DataFrame([summary])], ignore_index=True)
        df_out.loc[len(df_out) - 1, "ŚREDNIA_CENA_M2"] = avg if avg is not None else ""

        premium_cols = [
            "cena","cena_za_metr","metry","liczba_pokoi","pietro","rynek","rok_budowy",
            "material","wojewodztwo","powiat","gmina","miejscowosc","dzielnica","ulica","link",
            "ŚREDNIA_CENA_M2",
        ]
        existing = [c for c in premium_cols if c in df_out.columns]
        if existing:
            df_out = df_out[existing]

        try:
            df_out.to_excel(out_path, index=False)
        except Exception as e:
            messagebox.showerror("Błąd zapisu", f"Nie udało się zapisać pliku:\n{out_path}\n\n{e}")
            return

        if avg is not None and margin_pct > 0:
            corrected = avg * (1 - margin_pct / 100.0)
        else:
            corrected = avg

        col_avg = _find_col(self.df.columns, ["Średnia cena za m2 ( z bazy)", "Srednia cena za m2 ( z bazy)", "Średnia cena za m² (z bazy)"])
        col_avg_corr = _find_col(self.df.columns, ["Średnia skorygowana cena za m2", "Srednia skorygowana cena za m2"])
        col_stat = _find_col(self.df.columns, ["Statystyczna wartość nieruchomości", "Statystyczna wartosc nieruchomosci"])

        if col_avg is None:
            col_avg = VALUE_COLS[0]
            if col_avg not in self.df.columns:
                self.df[col_avg] = ""
        if col_avg_corr is None:
            col_avg_corr = VALUE_COLS[1]
            if col_avg_corr not in self.df.columns:
                self.df[col_avg_corr] = ""
        if col_stat is None:
            col_stat = VALUE_COLS[2]
            if col_stat not in self.df.columns:
                self.df[col_stat] = ""

        self.df.at[self.current_idx, col_avg] = avg if avg is not None else ""
        self.df.at[self.current_idx, col_avg_corr] = corrected if corrected is not None else ""
        stat_val = (area_val * corrected) if (area_val is not None and corrected is not None) else ""
        self.df.at[self.current_idx, col_stat] = stat_val

        # ✅ Zapis Excela tylko do arkusza 'raport' (bez kasowania innych arkuszy)
        try:
            if self.input_path and self.input_path.suffix.lower() in (".xlsx", ".xlsm"):
                _write_df_to_sheet_preserve(self.input_path, self.df, sheet_name=RAPORT_SHEET)
            elif self.input_path and self.input_path.suffix.lower() == ".csv":
                self.df.to_csv(self.input_path, index=False, encoding="utf-8-sig")
        except Exception as e:
            messagebox.showwarning(
                "Zapis raportu",
                f"Wyliczono wartości, ale nie udało się zapisać raportu:\n{self.input_path}\n\n{e}",
            )

        msg = [f"Zapisano dobrane rekordy do: {out_path}"]
        if avg is not None:
            msg.append("Średnia cena/m²: " + f"{avg:,.2f}".replace(",", " ").replace(".", ","))
        if corrected is not None and corrected != avg:
            msg.append(f"Średnia po obniżce ({margin_pct:.1f}%): " + f"{corrected:,.2f}".replace(",", " ").replace(".", ","))
        if isinstance(stat_val, (int, float)):
            msg.append("Statystyczna wartość: " + f"{stat_val:,.2f}".replace(",", " ").replace(".", ","))
        messagebox.showinfo("Zakończono", "\n".join(msg))

def main():
    app = App()
    app.mainloop()

if __name__ == "__main__":
    main()
