#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from __future__ import annotations

'''
automat.py — Wersja B (BDL + lokalna ludność + bezpieczny zapis arkusza 'raport')

FIX (17.12.2025+):
- ludnosc.csv jest wczytywane OK (logi), ale brak trafień wynikał z różnic w nazwach (pow./powiat, gmina miejska..., nawiasy)
- dodano kanonizację nazw jednostek (usuwa prefiksy/skrótowce/nawiasy)
- dodano fallback dopasowania po (woj + miejscowosc) + preferencja dzielnicy
- zapis XLSX: openpyxl, tylko arkusz 'raport' (bez kasowania innych arkuszy)
'''

from pathlib import Path
import sys
import unicodedata
import csv
import os
import datetime
import re
from typing import Optional, Dict, List, Tuple
from dataclasses import dataclass

import pandas as pd
import numpy as np

import tkinter as tk
from tkinter import ttk
import importlib.util
import requests
from openpyxl import load_workbook
from pathlib import Path

def import_local_automat():
    here = Path(__file__).resolve().parent
    p = here / "automat.py"
    spec = importlib.util.spec_from_file_location("automat", str(p))
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod

# =========================
# Helpers
# =========================

def _norm(s: str) -> str:
    return (s or "").strip().lower().replace(" ", "").replace("\xa0", "").replace("\t", "")

def _plain(x) -> str:
    """Bezpieczna normalizacja tekstu dla dowolnego typu (str/float/None/NaN)."""
    if x is None:
        return ""
    try:
        if isinstance(x, float) and np.isnan(x):
            return ""
    except Exception:
        pass

    s = str(x).strip().lower()
    s = "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))
    s = " ".join(s.split())
    return s

def _strip_parentheses(s: str) -> str:
    # usuń nawiasy i zawartość: "Gdańsk (miasto)" -> "Gdańsk"
    return re.sub(r"\([^)]*\)", " ", s).strip()

def _canon_admin(part: str, kind: str) -> str:
    """
    kind: woj/pow/gmi/mia/dzl
    Ujednolica teksty z raportu i csv:
    - usuwa nawiasy
    - usuwa znaki interpunkcyjne
    - usuwa słowa typu: powiat, pow., gmina, gm., woj., województwo, itd.
    """
    s = _plain(part)
    if not s:
        return ""
    s = _strip_parentheses(s)

    # zamień myślniki/slashe na spacje
    s = s.replace("-", " ").replace("/", " ")
    # wywal wszystko poza litery/cyfry/spacje
    s = re.sub(r"[^0-9a-z ]+", " ", s)
    s = " ".join(s.split())

    # tokeny do wywalenia
    drop_common = {
        "woj", "woj.", "wojewodztwo",
        "pow", "pow.", "powiat",
        "gmina", "gm", "gm.",
        "miasto", "m", "m.",
        "osiedle", "dzielnica",
        "miejska", "wiejska", "miejskowiejska", "miejsko", "wiejsko",
        "na", "prawach", "powiatu",
    }

    tokens = [t for t in s.split() if t not in drop_common]

    # czasem po usunięciu zostaje pusto – wtedy zostaw oryginalne (po plain)
    if not tokens:
        tokens = s.split()

    return " ".join(tokens).strip()

def _find_col(cols, candidates):
    norm_map = {_norm(c): c for c in cols}
    for cand in candidates:
        key = _norm(cand)
        if key in norm_map:
            return norm_map[key]
    for c in cols:
        if any(_norm(x) in _norm(c) for x in candidates):
            return c
    return None

def _trim_after_semicolon(val):
    if val is None:
        return ""
    try:
        if pd.isna(val):
            return ""
    except Exception:
        pass
    s = str(val)
    if ";" in s:
        s = s.split(";", 1)[0]
    return s.strip()

def _to_float_maybe(x):
    if x is None:
        return None
    try:
        if isinstance(x, float) and np.isnan(x):
            return None
    except Exception:
        pass

    s = str(x)
    for unit in ["m²", "m2", "zł/m²", "zł/m2", "zł"]:
        s = s.replace(unit, "")
    s = s.replace(" ", "").replace("\xa0", "").replace(",", ".")
    s = "".join(ch for ch in s if (ch.isdigit() or ch == "." or ch == "-"))
    try:
        return float(s) if s else None
    except Exception:
        return None


def _find_ludnosc_csv(baza_folder: Path, raport_path: Path, polska_path: Path) -> Path | None:
    candidates = [
        baza_folder / "ludnosc.csv",
        polska_path.parent / "ludnosc.csv",
        raport_path.parent / "ludnosc.csv",
        Path.cwd() / "ludnosc.csv",
        baza_folder / "Ludnosc.csv",
        polska_path.parent / "Ludnosc.csv",
        raport_path.parent / "Ludnosc.csv",
    ]
    for p in candidates:
        try:
            if p.exists():
                return p.resolve()
        except Exception:
            pass
    return None


# =========================
# Progi ludności
# =========================

POP_MARGIN_RULES = [
    (0,      6000,   25.0, 15.0),
    (6000,  20000,   20.0, 15.0),
    (20000,  50000,   20.0, 15.0),
    (50000,  200000,  15.0, 15.0),
    (200000, None,    10.0, 15.0),
]

def configure_margins_gui():
    root = tk.Tk()
    root.title("Ustawienia progów ludności")
    root.resizable(False, False)

    ttk.Label(root, text="Minimalna ludność").grid(row=0, column=0, padx=4, pady=4)
    ttk.Label(root, text="Maksymalna ludność").grid(row=0, column=1, padx=4, pady=4)
    ttk.Label(root, text="Pomiar brzegowy m²").grid(row=0, column=2, padx=4, pady=4)
    ttk.Label(root, text="% negocjacyjny").grid(row=0, column=3, padx=4, pady=4)

    entries_m2: list[ttk.Entry] = []
    entries_pct: list[ttk.Entry] = []

    def _fmt_pop(x):
        if x is None:
            return "∞"
        try:
            x = int(x)
        except Exception:
            return str(x)
        return f"{x:,}".replace(",", " ")

    for i, (low, high, m2, pct) in enumerate(POP_MARGIN_RULES, start=1):
        ttk.Label(root, text=_fmt_pop(low)).grid(row=i, column=0, padx=4, pady=2, sticky="e")
        ttk.Label(root, text=_fmt_pop(high)).grid(row=i, column=1, padx=4, pady=2, sticky="e")

        e_m2 = ttk.Entry(root, width=8, justify="right")
        e_m2.insert(0, str(m2))
        e_m2.grid(row=i, column=2, padx=4, pady=2)
        entries_m2.append(e_m2)

        e_pct = ttk.Entry(root, width=8, justify="right")
        e_pct.insert(0, str(pct))
        e_pct.grid(row=i, column=3, padx=4, pady=2)
        entries_pct.append(e_pct)

    result = {"ok": False, "rules": POP_MARGIN_RULES}

    def on_ok():
        new_rules = []
        for (low, high, default_m2, default_pct), e_m2, e_pct in zip(POP_MARGIN_RULES, entries_m2, entries_pct):
            raw_m2 = e_m2.get().strip().replace(" ", "").replace(",", ".")
            raw_pct = e_pct.get().strip().replace(" ", "").replace(",", ".")
            try:
                m2_val = float(raw_m2) if raw_m2 else float(default_m2)
            except Exception:
                m2_val = float(default_m2)
            try:
                pct_val = float(raw_pct) if raw_pct else float(default_pct)
            except Exception:
                pct_val = float(default_pct)
            new_rules.append((low, high, m2_val, pct_val))
        result["ok"] = True
        result["rules"] = new_rules
        root.destroy()

    def on_cancel():
        result["ok"] = False
        root.destroy()

    btn_frame = ttk.Frame(root)
    btn_frame.grid(row=len(POP_MARGIN_RULES) + 1, column=0, columnspan=4, pady=(8, 8))
    ttk.Button(btn_frame, text="Anuluj", command=on_cancel).pack(side="right", padx=4)
    ttk.Button(btn_frame, text="Start", command=on_ok).pack(side="right", padx=4)

    root.update_idletasks()
    w, h = root.winfo_width(), root.winfo_height()
    x = (root.winfo_screenwidth() - w) // 2
    y = (root.winfo_screenheight() - h) // 2
    root.geometry(f"{w}x{h}+{x}+{y}")
    root.mainloop()

    return result["rules"] if result["ok"] else None

def rules_for_population(pop):
    if pop is None:
        return float(POP_MARGIN_RULES[-1][2]), float(POP_MARGIN_RULES[-1][3])
    try:
        p = float(pop)
    except Exception:
        return float(POP_MARGIN_RULES[-1][2]), float(POP_MARGIN_RULES[-1][3])

    for low, high, m2, pct in POP_MARGIN_RULES:
        if p >= low and (high is None or p < high):
            return float(m2), float(pct)
    return float(POP_MARGIN_RULES[-1][2]), float(POP_MARGIN_RULES[-1][3])

def _eq_mask(df: pd.DataFrame, col_candidates, value: str) -> pd.Series:
    col = _find_col(df.columns, col_candidates)
    if col is None or not str(value).strip():
        return pd.Series(True, index=df.index)
    s = df[col].astype(str).str.strip().str.lower()
    v = str(value).strip().lower()
    return s == v


# =========================
# BDL / ludność
# =========================

BDL_BASE_URL = "https://bdl.stat.gov.pl/api/v1"
BDL_API_KEY_DEFAULT = "c804c054-f519-45b3-38f3-08de375a07dc"

_BDL_POP_VAR_ID: str | None = None
_BDL_POP_VAR_ID_NOT_FOUND = "__NOT_FOUND__"

def _bdl_headers() -> dict:
    api_key = os.getenv("BDL_API_KEY") or os.getenv("GUS_BDL_API_KEY") or BDL_API_KEY_DEFAULT
    if not api_key:
        return {}
    return {"X-ClientId": api_key, "Accept": "application/json"}

def _pick_latest_year():
    return datetime.date.today().year - 1


class PopulationResolver:
    def __init__(self, local_csv: Path | None, api_cache_csv: Path | None, use_api: bool = True):
        self.local_csv = local_csv
        self.api_cache_csv = api_cache_csv
        self.use_api = bool(use_api)
        self._local: Dict[str, float] = {}
        self._api_cache: Dict[str, float] = {}
        self._dirty = False
        self._debug_miss = 0
        self._load_local()
        self._load_api_cache()

    def _make_key(self, woj: str = "", powiat: str = "", gmina: str = "", miejscowosc: str = "", dzielnica: str = "") -> str:
        w = _canon_admin(woj, "woj")
        p = _canon_admin(powiat, "pow")
        g = _canon_admin(gmina, "gmi")
        m = _canon_admin(miejscowosc, "mia")
        d = _canon_admin(dzielnica, "dzl")
        return "|".join([w, p, g, m, d])

    def _split_key(self, key: str) -> Tuple[str, str, str, str, str]:
        parts = (key.split("|") + ["", "", "", "", ""])[:5]
        return parts[0], parts[1], parts[2], parts[3], parts[4]

    def _candidate_keys(self, woj: str, powiat: str, gmina: str, miejscowosc: str, dzielnica: str) -> List[str]:
        # podstawowa hierarchia
        keys = [
            self._make_key(woj, powiat, gmina, miejscowosc, dzielnica),
            self._make_key(woj, powiat, gmina, miejscowosc, ""),
            self._make_key(woj, powiat, gmina, "", ""),
            self._make_key(woj, powiat, "", "", ""),
            self._make_key(woj, "", "", "", ""),
        ]

        # dodatkowe ścieżki gdy raport ma puste powiat/gmina, a csv ma wypełnione:
        keys += [
            self._make_key(woj, "", gmina, miejscowosc, dzielnica),
            self._make_key(woj, "", gmina, miejscowosc, ""),
            self._make_key(woj, "", gmina, "", ""),
            self._make_key(woj, powiat, "", miejscowosc, dzielnica),
            self._make_key(woj, powiat, "", miejscowosc, ""),
            self._make_key(woj, "", "", miejscowosc, dzielnica),
            self._make_key(woj, "", "", miejscowosc, ""),
        ]

        out, seen = [], set()
        for k in keys:
            if not k or k in seen:
                continue
            seen.add(k)
            out.append(k)
        return out

    def _read_local_csv_any_sep(self, path: Path) -> pd.DataFrame:
        for sep in [";", ",", "\t"]:
            try:
                return pd.read_csv(path, sep=sep, dtype=str, encoding="utf-8-sig", engine="python")
            except Exception:
                continue
        return pd.read_csv(path, sep=None, dtype=str, encoding="utf-8-sig", engine="python")

    def _load_local(self):
        if not self.local_csv:
            print("[PopulationResolver] local_csv=None (nie podano ścieżki).")
            return
        if not self.local_csv.exists():
            print(f"[PopulationResolver] local ludnosc.csv: NIE ISTNIEJE -> {self.local_csv}")
            return

        print(f"[PopulationResolver] Wczytuję local ludnosc.csv -> {self.local_csv}")

        try:
            df = self._read_local_csv_any_sep(self.local_csv)
            print(f"[PopulationResolver] local rows={len(df)} cols={list(df.columns)}")

            col_woj = _find_col(df.columns, ["Wojewodztwo", "Województwo"])
            col_pow = _find_col(df.columns, ["Powiat"])
            col_gmi = _find_col(df.columns, ["Gmina"])
            col_mia = _find_col(df.columns, ["Miejscowosc", "Miejscowość", "Miasto"])
            col_dzl = _find_col(df.columns, ["Dzielnica", "Osiedle"])
            col_pop = _find_col(df.columns, ["ludnosc", "Ludnosc", "Liczba mieszkancow", "Liczba mieszkańców", "population"])

            print(f"[PopulationResolver] map cols: woj={col_woj} pow={col_pow} gmi={col_gmi} mia={col_mia} dzl={col_dzl} pop={col_pop}")

            if not col_pop:
                print("[PopulationResolver] local ludnosc.csv: brak kolumny ludnosc/population -> nie użyję pliku.")
                return

            loaded = 0
            for _, r in df.iterrows():
                pop_f = _to_float_maybe(r.get(col_pop))
                if pop_f is None:
                    continue

                woj = r.get(col_woj, "") if col_woj else ""
                powiat = r.get(col_pow, "") if col_pow else ""
                gmina = r.get(col_gmi, "") if col_gmi else ""
                miejsc = r.get(col_mia, "") if col_mia else ""
                dziel = r.get(col_dzl, "") if col_dzl else ""

                key = self._make_key(woj, powiat, gmina, miejsc, dziel)
                if key:
                    self._local[key] = float(pop_f)
                    loaded += 1

            print(f"[PopulationResolver] local loaded keys={loaded} (unikalne={len(self._local)})")

        except Exception as e:
            print(f"[PopulationResolver] Nie udało się wczytać local ludnosc.csv: {e}")

    def _load_api_cache(self):
        if not self.api_cache_csv or not self.api_cache_csv.exists():
            return
        try:
            with self.api_cache_csv.open("r", encoding="utf-8-sig", newline="") as f:
                rd = csv.DictReader(f)
                for row in rd:
                    pop = _to_float_maybe(row.get("population", ""))
                    if pop is None:
                        continue
                    key = row.get("key") or self._make_key(
                        row.get("woj", ""), row.get("powiat", ""), row.get("gmina", ""),
                        row.get("miejscowosc", ""), row.get("dzielnica", "")
                    )
                    if key:
                        self._api_cache[key] = float(pop)
        except Exception as e:
            print(f"[PopulationResolver] Nie udało się wczytać cache API: {e}")

    def _save_api_cache(self):
        if not self._dirty or not self.api_cache_csv:
            return
        try:
            self.api_cache_csv.parent.mkdir(parents=True, exist_ok=True)
            with self.api_cache_csv.open("w", encoding="utf-8-sig", newline="") as f:
                fieldnames = ["key", "woj", "powiat", "gmina", "miejscowosc", "dzielnica", "population"]
                wr = csv.DictWriter(f, fieldnames=fieldnames)
                wr.writeheader()
                for key, pop in self._api_cache.items():
                    parts = (key.split("|") + ["", "", "", "", ""])[:5]
                    woj, pow, gmi, mia, dzl = parts
                    wr.writerow({
                        "key": key,
                        "woj": woj,
                        "powiat": pow,
                        "gmina": gmi,
                        "miejscowosc": mia,
                        "dzielnica": dzl,
                        "population": pop,
                    })
            self._dirty = False
        except Exception as e:
            print(f"[PopulationResolver] Nie udało się zapisać cache API: {e}")

    def _get_population_var_id(self) -> str | None:
        global _BDL_POP_VAR_ID

        if _BDL_POP_VAR_ID == _BDL_POP_VAR_ID_NOT_FOUND:
            return None
        if _BDL_POP_VAR_ID:
            return _BDL_POP_VAR_ID

        headers = _bdl_headers()
        if not headers:
            return None

        try:
            url = f"{BDL_BASE_URL}/variables"
            params = {"name": "ludność ogółem", "page-size": 50, "format": "json"}
            r = requests.get(url, headers=headers, params=params, timeout=15)
            if r.status_code == 200:
                data = r.json()
                for v in data.get("results", []):
                    name = (v.get("name") or "").lower()
                    if "ludność ogółem" in name or "ludnosc ogolem" in name or "population total" in name:
                        _BDL_POP_VAR_ID = str(v.get("id"))
                        print(f"[PopulationResolver] Zmienna ludności: id={_BDL_POP_VAR_ID} ({name})")
                        return _BDL_POP_VAR_ID
        except Exception:
            pass

        print("[PopulationResolver] Nie znalazłem zmiennej 'ludność ogółem' w BDL (cache).")
        _BDL_POP_VAR_ID = _BDL_POP_VAR_ID_NOT_FOUND
        return None

    def _fetch_population_from_api(self, woj: str, powiat: str, gmina: str, miejscowosc: str) -> Optional[float]:
        headers = _bdl_headers()
        if not headers:
            return None

        name_search = miejscowosc or gmina
        if not name_search:
            return None

        try:
            url_units = f"{BDL_BASE_URL}/units"
            params_units = {"name": name_search, "level": "6", "page-size": 50, "format": "json"}
            ru = requests.get(url_units, headers=headers, params=params_units, timeout=15)
            if ru.status_code != 200:
                return None
            ju = ru.json()
            units = ju.get("results", []) or []
            if not units:
                return None

            def score(u):
                nm = _plain(u.get("name") or "")
                sc = 0
                if _plain(name_search) == nm:
                    sc += 5
                elif _plain(name_search) in nm:
                    sc += 3
                if powiat and _plain(powiat) in nm:
                    sc += 1
                if woj and _plain(woj) in nm:
                    sc += 1
                return sc

            units.sort(key=score, reverse=True)
            unit_id = units[0].get("id")
            if not unit_id:
                return None
        except Exception:
            return None

        var_id = self._get_population_var_id()
        if not var_id:
            return None

        year = _pick_latest_year()
        try:
            url_data = f"{BDL_BASE_URL}/data/by-unit/{unit_id}"
            params_data = {"var-id": var_id, "year": str(year), "format": "json"}
            rd = requests.get(url_data, headers=headers, params=params_data, timeout=20)
            if rd.status_code != 200:
                return None

            jd = rd.json()
            results = jd.get("results") or []
            if not results:
                return None

            vals = results[0].get("values") or []
            for v in vals:
                raw = v[0] if isinstance(v, list) and len(v) >= 1 else v
                pop = _to_float_maybe(raw)
                if pop is not None:
                    return float(pop)
        except Exception:
            return None

        return None

    def _fallback_by_woj_mia(self, woj: str, miejscowosc: str, dzielnica: str) -> Optional[float]:
        """
        Jeżeli pełne klucze nie trafiają (różnice w pow/gmi), spróbuj:
        - dopasować po (woj + miejscowosc)
        - jeśli dzielnica podana, preferuj rekordy z tą dzielnicą
        """
        woj_c = _canon_admin(woj, "woj")
        mia_c = _canon_admin(miejscowosc, "mia")
        dzl_c = _canon_admin(dzielnica, "dzl")

        if not woj_c or not mia_c:
            return None

        best_with_dzl = None
        best_any = None

        for key, pop in self._local.items():
            w, p, g, m, d = self._split_key(key)
            if w != woj_c or m != mia_c:
                continue
            if dzl_c and d == dzl_c:
                # preferuj dokładną dzielnicę; jeśli kilka, bierz największą (bezpiecznie)
                best_with_dzl = pop if (best_with_dzl is None or pop > best_with_dzl) else best_with_dzl
            else:
                best_any = pop if (best_any is None or pop > best_any) else best_any

        return best_with_dzl if best_with_dzl is not None else best_any

    def get_population(self, woj: str, powiat: str, gmina: str, miejscowosc: str, dzielnica: str) -> Optional[float]:
        # 1) local/cache: po kandydatach
        for key in self._candidate_keys(woj, powiat, gmina, miejscowosc, dzielnica):
            if key in self._local:
                return self._local[key]
            if key in self._api_cache:
                return self._api_cache[key]

        # 2) fallback: woj + miejscowosc (często raport ma inne pow/gmi niż csv)
        pop = self._fallback_by_woj_mia(woj, miejscowosc, dzielnica)
        if pop is not None:
            return float(pop)

        # 3) API
        if self.use_api:
            pop = self._fetch_population_from_api(woj, powiat, gmina, miejscowosc)
            if pop is not None:
                key4 = self._make_key(woj, powiat, gmina, miejscowosc, "")
                self._api_cache[key4] = float(pop)
                self._dirty = True
                self._save_api_cache()
                return float(pop)

        # mała diagnostyka: pokaż pierwsze 3 nietrafienia (żeby nie spamować)
        if self._debug_miss < 3:
            self._debug_miss += 1
            print("[PopulationResolver][MISS] szukałem dla:")
            print("  woj=", woj, "pow=", powiat, "gmi=", gmina, "mia=", miejscowosc, "dzl=", dzielnica)
            print("  canon key=", self._make_key(woj, powiat, gmina, miejscowosc, dzielnica))

        return None


# =========================
# Bezpieczny zapis XLSX (TYLKO arkusz 'raport')
# =========================

def _pick_report_sheet_name(xlsx_path: Path, preferred: str = "raport") -> str:
    wb = load_workbook(xlsx_path)
    if preferred in wb.sheetnames:
        return preferred
    return wb.sheetnames[0] if wb.sheetnames else preferred

def save_report_sheet_only(xlsx_path: Path, df_report: pd.DataFrame, sheet_name: str = "raport") -> None:
    wb = load_workbook(xlsx_path)

    if sheet_name not in wb.sheetnames:
        if wb.sheetnames:
            sheet_name = wb.sheetnames[0]
        else:
            wb.create_sheet(sheet_name)

    ws = wb[sheet_name]

    if ws.max_row >= 1:
        ws.delete_rows(1, ws.max_row)

    for c, name in enumerate(df_report.columns.tolist(), start=1):
        ws.cell(row=1, column=c, value=name)

    for r_idx, row in enumerate(df_report.values.tolist(), start=2):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    wb.save(xlsx_path)


# =========================
# Core: przetwarzanie wiersza
# =========================

def _bucket_for_population(pop: float | None) -> tuple[float | None, float | None]:
    """Zwraca (low, high) dla progu ludności wg POP_MARGIN_RULES."""
    if pop is None:
        return (None, None)
    try:
        p = float(pop)
    except Exception:
        return (None, None)

    for low, high, _, _ in POP_MARGIN_RULES:
        if p >= low and (high is None or p < high):
            return (float(low), float(high) if high is not None else None)

    # fallback: ostatni próg
    low, high, _, _ = POP_MARGIN_RULES[-1]
    return (float(low), float(high) if high is not None else None)

def _pop_in_bucket(pop: float | None, low: float | None, high: float | None) -> bool:
    if low is None and high is None:
        return True
    if pop is None:
        return False
    try:
        p = float(pop)
    except Exception:
        return False
    if high is None:
        return p >= float(low)
    return p >= float(low) and p < float(high)

@dataclass
class PolskaIndex:
    df: pd.DataFrame
    col_area: str
    col_price: str

    # kolumny adresowe w Polsce.xlsx (oryginalne)
    col_woj: str | None
    col_pow: str | None
    col_gmi: str | None
    col_mia: str | None
    col_dzl: str | None

    # kolumny kanoniczne
    c_woj: str | None
    c_pow: str | None
    c_gmi: str | None
    c_mia: str | None
    c_dzl: str | None

    # mapy miejscowości -> przykładowa nazwa (żeby móc pytać PopulationResolver sensownym tekstem)
    by_gmina: Dict[Tuple[str, str, str], Dict[str, str]]
    by_powiat: Dict[Tuple[str, str], Dict[str, str]]
    by_woj: Dict[str, Dict[str, str]]

def build_polska_index(df_pl: pd.DataFrame, col_area_pl: str, col_price_pl: str) -> PolskaIndex:
    # kolumny administracyjne
    col_woj = _find_col(df_pl.columns, ["wojewodztwo", "województwo", "woj"])
    col_pow = _find_col(df_pl.columns, ["powiat"])
    col_gmi = _find_col(df_pl.columns, ["gmina"])
    col_mia = _find_col(df_pl.columns, ["miejscowosc", "miejscowość", "miasto"])
    col_dzl = _find_col(df_pl.columns, ["dzielnica", "osiedle"])

    # numeryczne metry/cena (jednorazowo – szybciej per wiersz raportu)
    if "_area_num" not in df_pl.columns:
        df_pl["_area_num"] = df_pl[col_area_pl].map(_to_float_maybe)
    if "_price_num" not in df_pl.columns:
        df_pl["_price_num"] = df_pl[col_price_pl].map(_to_float_maybe)

    # kanonizacja tekstów do porównań
    c_woj = c_pow = c_gmi = c_mia = c_dzl = None
    if col_woj:
        c_woj = "_woj_c"
        df_pl[c_woj] = df_pl[col_woj].map(lambda x: _canon_admin(x, "woj"))
    if col_pow:
        c_pow = "_pow_c"
        df_pl[c_pow] = df_pl[col_pow].map(lambda x: _canon_admin(x, "pow"))
    if col_gmi:
        c_gmi = "_gmi_c"
        df_pl[c_gmi] = df_pl[col_gmi].map(lambda x: _canon_admin(x, "gmi"))
    if col_mia:
        c_mia = "_mia_c"
        df_pl[c_mia] = df_pl[col_mia].map(lambda x: _canon_admin(x, "mia"))
    if col_dzl:
        c_dzl = "_dzl_c"
        df_pl[c_dzl] = df_pl[col_dzl].map(lambda x: _canon_admin(x, "dzl"))

    # mapy miejscowości po gminie/powiecie/woj (na bazie tego co w ogóle jest w Polska.xlsx)
    by_gmina: Dict[Tuple[str, str, str], Dict[str, str]] = {}
    by_powiat: Dict[Tuple[str, str], Dict[str, str]] = {}
    by_woj: Dict[str, Dict[str, str]] = {}

    # brak miejscowości -> brak indeksu (zostaną puste mapy, a fallback zadziała tylko na miejscowości)
    if c_woj and c_mia:
        # województwo
        for w, gdf in df_pl.groupby(c_woj, dropna=False):
            if not w:
                continue
            mp: Dict[str, str] = {}
            if col_mia and c_mia:
                for mia_c, sub in gdf.groupby(c_mia, dropna=False):
                    if not mia_c:
                        continue
                    # przykładowa oryginalna nazwa (z diakrytykami)
                    try:
                        ex = sub[col_mia].dropna().iloc[0]
                        mp[mia_c] = str(ex) if pd.notna(ex) else str(mia_c)
                    except Exception:
                        mp[mia_c] = str(mia_c)
            by_woj[str(w)] = mp

        # powiat
        if c_pow:
            for (w, p), gdf in df_pl.groupby([c_woj, c_pow], dropna=False):
                if not w or not p:
                    continue
                mp: Dict[str, str] = {}
                if col_mia:
                    for mia_c, sub in gdf.groupby(c_mia, dropna=False):
                        if not mia_c:
                            continue
                        try:
                            ex = sub[col_mia].dropna().iloc[0]
                            mp[mia_c] = str(ex) if pd.notna(ex) else str(mia_c)
                        except Exception:
                            mp[mia_c] = str(mia_c)
                by_powiat[(str(w), str(p))] = mp

        # gmina
        if c_pow and c_gmi:
            for (w, p, g), gdf in df_pl.groupby([c_woj, c_pow, c_gmi], dropna=False):
                if not w or not p or not g:
                    continue
                mp: Dict[str, str] = {}
                if col_mia:
                    for mia_c, sub in gdf.groupby(c_mia, dropna=False):
                        if not mia_c:
                            continue
                        try:
                            ex = sub[col_mia].dropna().iloc[0]
                            mp[mia_c] = str(ex) if pd.notna(ex) else str(mia_c)
                        except Exception:
                            mp[mia_c] = str(mia_c)
                by_gmina[(str(w), str(p), str(g))] = mp

    return PolskaIndex(
        df=df_pl,
        col_area=col_area_pl,
        col_price=col_price_pl,
        col_woj=col_woj,
        col_pow=col_pow,
        col_gmi=col_gmi,
        col_mia=col_mia,
        col_dzl=col_dzl,
        c_woj=c_woj,
        c_pow=c_pow,
        c_gmi=c_gmi,
        c_mia=c_mia,
        c_dzl=c_dzl,
        by_gmina=by_gmina,
        by_powiat=by_powiat,
        by_woj=by_woj,
    )

def _mask_eq_canon(df: pd.DataFrame, canon_col: str | None, value_canon: str) -> pd.Series:
    if canon_col is None or not value_canon:
        return pd.Series(True, index=df.index)
    return df[canon_col].astype(str) == str(value_canon)

def _filter_miejscowosci_by_bucket(
    candidates: Dict[str, str],
    bucket_low: float | None,
    bucket_high: float | None,
    pop_resolver: PopulationResolver | None,
    woj_raw: str,
    pow_raw: str,
    gmi_raw: str,
    scope: str,
    pop_cache: Dict[Tuple[str, str], float | None],
) -> List[str]:
    """Zwraca listę kanonicznych nazw miejscowości w zadanym progu ludności.
    Jeżeli brakuje ludności dla wszystkich, zwróci pustą listę (wtedy wyżej można zrobić fallback)."""
    if not candidates:
        return []
    if bucket_low is None and bucket_high is None:
        return list(candidates.keys())

    out: List[str] = []
    for mia_c, mia_original in candidates.items():
        cache_key = (scope, mia_c)
        if cache_key in pop_cache:
            pop = pop_cache[cache_key]
        else:
            pop = None
            if pop_resolver is not None:
                pop = pop_resolver.get_population(woj_raw, pow_raw, gmi_raw, mia_original, "")
            pop_cache[cache_key] = pop
        if _pop_in_bucket(pop, bucket_low, bucket_high):
            out.append(mia_c)
    return out

def _select_comparables(
    pl: PolskaIndex,
    woj_c: str,
    pow_c: str,
    gmi_c: str,
    mia_c: str,
    dzl_c: str,
    woj_raw: str,
    pow_raw: str,
    gmi_raw: str,
    min_hits: int,
    bucket_low: float | None,
    bucket_high: float | None,
    pop_resolver: PopulationResolver | None,
    low_area: float,
    high_area: float,
) -> Tuple[pd.DataFrame, str]:
    df = pl.df

    # baza: zakres metrażu + musi mieć cenę
    base = (df["_area_num"].notna()) & (df["_area_num"] >= low_area) & (df["_area_num"] <= high_area) & df["_price_num"].notna()

    pop_cache: Dict[Tuple[str, str], float | None] = {}

    def _take(mask: pd.Series, label: str) -> Tuple[pd.DataFrame, str]:
        sel = df[mask].copy()
        # tylko z ceną (już jest), ale zabezpiecz
        sel = sel[sel["_price_num"].notna()].copy()
        return sel, label

    # 1) DZIELNICA (jeśli podana)
    if dzl_c:
        mask = base.copy()
        mask &= _mask_eq_canon(df, pl.c_woj, woj_c)
        mask &= _mask_eq_canon(df, pl.c_mia, mia_c)
        mask &= _mask_eq_canon(df, pl.c_dzl, dzl_c)
        sel, label = _take(mask, "dzielnica")
        if len(sel.index) >= min_hits:
            return sel, label

    # 2) MIEJSCOWOŚĆ
    if mia_c:
        mask = base.copy()
        mask &= _mask_eq_canon(df, pl.c_woj, woj_c)
        mask &= _mask_eq_canon(df, pl.c_mia, mia_c)
        sel, label = _take(mask, "miejscowosc")
        if len(sel.index) >= min_hits:
            return sel, label

    # 3) GMINA (miejscowości w tym samym progu ludności)
    if gmi_c and pl.c_mia and pl.by_gmina:
        candidates = pl.by_gmina.get((woj_c, pow_c, gmi_c), {})
        bucket_mias = _filter_miejscowosci_by_bucket(
            candidates, bucket_low, bucket_high, pop_resolver,
            woj_raw=woj_raw, pow_raw=pow_raw, gmi_raw=gmi_raw,
            scope="gmina", pop_cache=pop_cache
        )
        if not bucket_mias:
            # fallback: jak nie mamy ludności (lub brak danych), weź wszystkie miejscowości z tej gminy
            bucket_mias = list(candidates.keys())

        if bucket_mias:
            mask = base.copy()
            mask &= _mask_eq_canon(df, pl.c_woj, woj_c)
            mask &= _mask_eq_canon(df, pl.c_pow, pow_c)
            mask &= _mask_eq_canon(df, pl.c_gmi, gmi_c)
            mask &= df[pl.c_mia].isin(bucket_mias)
            sel, label = _take(mask, "gmina(pop)")
            if len(sel.index) >= min_hits:
                return sel, label

    # 4) POWIAT (miejscowości w tym samym progu ludności)
    if pow_c and pl.c_mia and pl.by_powiat:
        candidates = pl.by_powiat.get((woj_c, pow_c), {})
        bucket_mias = _filter_miejscowosci_by_bucket(
            candidates, bucket_low, bucket_high, pop_resolver,
            woj_raw=woj_raw, pow_raw=pow_raw, gmi_raw="",
            scope="powiat", pop_cache=pop_cache
        )
        if not bucket_mias:
            bucket_mias = list(candidates.keys())

        if bucket_mias:
            mask = base.copy()
            mask &= _mask_eq_canon(df, pl.c_woj, woj_c)
            mask &= _mask_eq_canon(df, pl.c_pow, pow_c)
            mask &= df[pl.c_mia].isin(bucket_mias)
            sel, label = _take(mask, "powiat(pop)")
            if len(sel.index) >= min_hits:
                return sel, label

    # 5) WOJEWÓDZTWO (miejscowości w tym samym progu ludności)
    #    Specjalnie dla MAZOWIECKIEGO: zamiast przeszukiwać całe mazowieckie, przeszukuj WOJ. SĄSIEDNIE (bez mazowieckiego)
    #    i zbierz WSZYSTKIE ogłoszenia z tych województw do wyliczeń.
    if woj_c and pl.c_mia and pl.by_woj:
        if woj_c == "mazowieckie":
            neighbors = [
                "lodzkie",
                "kujawsko pomorskie",
                "warminsko mazurskie",
                "podlaskie",
                "lubelskie",
                "swietokrzyskie",
            ]

            parts = []
            for w2 in neighbors:
                candidates = pl.by_woj.get(w2, {})
                if not candidates:
                    continue

                bucket_mias = _filter_miejscowosci_by_bucket(
                    candidates, bucket_low, bucket_high, pop_resolver,
                    woj_raw=w2, pow_raw="", gmi_raw="",
                    scope=f"woj_sas:{w2}", pop_cache=pop_cache
                )
                if not bucket_mias:
                    bucket_mias = list(candidates.keys())

                if not bucket_mias:
                    continue

                mask = base.copy()
                mask &= _mask_eq_canon(df, pl.c_woj, w2)
                mask &= df[pl.c_mia].isin(bucket_mias)
                sel_part, _ = _take(mask, f"woj_sas:{w2}")
                if not sel_part.empty:
                    parts.append(sel_part)

            if parts:
                sel = pd.concat(parts, axis=0, ignore_index=False)
                # usuń duplikaty (na wszelki wypadek)
                sel = sel.loc[~sel.index.duplicated(keep="first")].copy()
                # UWAGA: tu NIE przerywamy po pierwszych 5 – zbieramy pełny zbiór z sąsiadów do wyliczeń.
                if len(sel.index) >= min_hits:
                    return sel, "woj_sasiednie(pop)"
                # jeśli < min_hits, i tak zwrócimy to jako najlepszy szeroki zestaw (zamiast pustego)
                return sel, "woj_sasiednie(pop)_malo"

        # standard: dla innych województw przeszukaj własne woj.
        candidates = pl.by_woj.get(woj_c, {})
        bucket_mias = _filter_miejscowosci_by_bucket(
            candidates, bucket_low, bucket_high, pop_resolver,
            woj_raw=woj_raw, pow_raw="", gmi_raw="",
            scope="woj", pop_cache=pop_cache
        )
        if not bucket_mias:
            bucket_mias = list(candidates.keys())

        if bucket_mias:
            mask = base.copy()
            mask &= _mask_eq_canon(df, pl.c_woj, woj_c)
            mask &= df[pl.c_mia].isin(bucket_mias)
            sel, label = _take(mask, "woj(pop)")
            if len(sel.index) >= min_hits:
                return sel, label

    # jeśli nadal < min_hits, zwróć najlepsze co mamy (ostatni znaleziony) albo puste
    # spróbuj chociaż na woj+miejscowość bez progu (jeśli progi odfiltrowały wszystko)
    if woj_c and mia_c:
        mask = base.copy()
        mask &= _mask_eq_canon(df, pl.c_woj, woj_c)
        mask &= _mask_eq_canon(df, pl.c_mia, mia_c)
        sel, label = _take(mask, "miejscowosc(fallback)")
        if not sel.empty:
            return sel, label

    return df.iloc[0:0].copy(), "brak"

def _process_row(
    df_raport: pd.DataFrame,
    idx: int,
    pl: PolskaIndex,
    margin_m2_default: float,
    margin_pct_default: float,
    pop_resolver: PopulationResolver | None,
    min_hits: int = 5,
) -> None:
    row = df_raport.iloc[idx]

    kw_col = _find_col(df_raport.columns, ["Nr KW", "nr_kw", "nrksiegi", "nr księgi", "nr_ksiegi", "numer księgi"])
    kw_value = (str(row[kw_col]).strip() if (kw_col and pd.notna(row[kw_col]) and str(row[kw_col]).strip()) else f"WIERSZ_{idx+1}")

    area_col = _find_col(df_raport.columns, ["Obszar", "metry", "powierzchnia"])
    area_val = _to_float_maybe(_trim_after_semicolon(row[area_col])) if area_col else None
    if area_val is None:
        print(f"[Automat] Wiersz {idx+1}: brak obszaru – pomijam.")
        return

    def _get(cands):
        c = _find_col(df_raport.columns, cands)
        return _trim_after_semicolon(row[c]) if c else ""

    woj_r = _get(["Województwo", "Wojewodztwo", "wojewodztwo", "woj"])
    pow_r = _get(["Powiat"])
    gmi_r = _get(["Gmina"])
    mia_r = _get(["Miejscowość", "Miejscowosc", "Miasto"])
    dzl_r = _get(["Dzielnica", "Osiedle"])

    woj_c = _canon_admin(woj_r, "woj")
    pow_c = _canon_admin(pow_r, "pow")
    gmi_c = _canon_admin(gmi_r, "gmi")
    mia_c = _canon_admin(mia_r, "mia")
    dzl_c = _canon_admin(dzl_r, "dzl")

    # =========================
    # TRYB STRICT (adres 100% wymagany)
    # =========================
    STRICT_MSG = "BRAK LUB NIEPEŁNY ADRESU – WPISZ ADRES MANUALNIE"

    mean_col = _find_col(df_raport.columns, ["Średnia cena za m2 ( z bazy)", "Srednia cena za m2 ( z bazy)", "Średnia cena za m² ( z bazy)"])
    corr_col = _find_col(df_raport.columns, ["Średnia skorygowana cena za m2", "Srednia skorygowana cena za m2"])
    val_col = _find_col(df_raport.columns, ["Statystyczna wartość nieruchomości", "Statystyczna wartosc nieruchomosci"])

    if mean_col is None:
        mean_col = "Średnia cena za m2 ( z bazy)"
        if mean_col not in df_raport.columns:
            df_raport[mean_col] = ""
    if corr_col is None:
        corr_col = "Średnia skorygowana cena za m2"
        if corr_col not in df_raport.columns:
            df_raport[corr_col] = ""
    if val_col is None:
        val_col = "Statystyczna wartość nieruchomości"
        if val_col not in df_raport.columns:
            df_raport[val_col] = ""

    # Minimalne dane do pracy: woj + miejscowość
    if not woj_c or not mia_c:
        df_raport.at[idx, mean_col] = STRICT_MSG
        df_raport.at[idx, corr_col] = STRICT_MSG
        df_raport.at[idx, val_col] = STRICT_MSG
        print(f"[Automat] {kw_value}: {STRICT_MSG} (woj='{woj_r}', mia='{mia_r}')")
        return

    # ludność + progi (m2 oraz %)
    pop_target = pop_resolver.get_population(woj_r, pow_r, gmi_r, mia_r, dzl_r) if pop_resolver else None
    bucket_low, bucket_high = _bucket_for_population(pop_target)

    if pop_target is None:
        margin_m2_row, margin_pct_row = float(margin_m2_default), float(margin_pct_default)
    else:
        margin_m2_row, margin_pct_row = rules_for_population(pop_target)

    delta = abs(float(margin_m2_row or 0.0))
    low_area, high_area = max(0.0, float(area_val) - delta), float(area_val) + delta

    df_sel, stage = _select_comparables(
        pl=pl,
        woj_c=woj_c,
        pow_c=pow_c,
        gmi_c=gmi_c,
        mia_c=mia_c,
        dzl_c=dzl_c,
        woj_raw=woj_r,
        pow_raw=pow_r,
        gmi_raw=gmi_r,
        min_hits=int(min_hits),
        bucket_low=bucket_low,
        bucket_high=bucket_high,
        pop_resolver=pop_resolver,
        low_area=low_area,
        high_area=high_area,
    )

    if df_sel.empty:
        msg = "BRAK OGŁOSZEŃ W BAZIE DLA TEGO ZAKRESU"
        df_raport.at[idx, mean_col] = msg
        df_raport.at[idx, corr_col] = msg
        df_raport.at[idx, val_col] = msg
        print(f"[Automat] {kw_value}: {msg} (zakres {low_area:.2f}-{high_area:.2f} m², stage={stage})")
        return

    # usuwanie odstających (bez ryzyka, że spadnie poniżej min_hits)
    prices = df_sel["_price_num"].astype(float)
    if len(prices) >= max(10, int(min_hits)):
        q1 = np.nanpercentile(prices, 25)
        q3 = np.nanpercentile(prices, 75)
        iqr = q3 - q1
        lo = q1 - 1.5 * iqr
        hi = q3 + 1.5 * iqr
        df_filtered = df_sel[(prices >= lo) & (prices <= hi)].copy()
        if len(df_filtered.index) >= int(min_hits):
            df_sel = df_filtered
            prices = df_sel["_price_num"].astype(float)

    mean_price = float(np.nanmean(prices))

    # =========================
    # WYNIKI – ZAOKRĄGLENIE DO 2 MIEJSC
    # =========================
    mean_price_rounded = round(float(mean_price), 2)
    df_raport.at[idx, mean_col] = mean_price_rounded

    corrected_price = mean_price_rounded * (1.0 - float(margin_pct_row or 0.0) / 100.0)
    corrected_price_rounded = round(float(corrected_price), 2)
    df_raport.at[idx, corr_col] = corrected_price_rounded

    value = corrected_price_rounded * float(area_val)
    df_raport.at[idx, val_col] = round(float(value), 2)

    # log
    pop_txt = f"{int(pop_target):,}".replace(",", " ") if isinstance(pop_target, (int, float)) else "?"
    bucket_txt = f"{bucket_low}-{bucket_high if bucket_high is not None else '∞'}" if bucket_low is not None else "?"
    print(f"[Automat] {kw_value}: stage={stage} | hits={len(df_sel)} | pop={pop_txt} | bucket={bucket_txt} | mean={mean_price:.2f} | corr={corrected_price:.2f} | value={value:.2f}.")


# =========================
# MAIN

# =========================

def main(argv=None) -> int:
    global POP_MARGIN_RULES

    if argv is None:
        argv = sys.argv

    if len(argv) < 3:
        print("Użycie: automat.py RAPORT_PATH BAZA_FOLDER")
        return 1

    raport_path = Path(argv[1]).resolve()
    baza_folder = Path(argv[2]).resolve()

    if not raport_path.exists():
        print(f"[BŁĄD] Nie znaleziono raportu: {raport_path}")
        return 1

    polska_path = baza_folder / "Polska.xlsx"
    if not polska_path.exists():
        print(f"[BŁĄD] Nie znaleziono Polska.xlsx w folderze: {baza_folder}")
        return 1

    margin_m2_default = 15.0
    margin_pct_default = 15.0

    try:
        new_rules = configure_margins_gui()
    except Exception as e:
        print(f"[Automat] Błąd GUI progów ludności: {e}")
        new_rules = POP_MARGIN_RULES

    if new_rules is None:
        print("[Automat] Przerwano (Anuluj w oknie progów ludności).")
        return 1
    POP_MARGIN_RULES = new_rules

    try:
        if len(POP_MARGIN_RULES) >= 3:
            margin_m2_default = float(POP_MARGIN_RULES[2][2])
            margin_pct_default = float(POP_MARGIN_RULES[2][3])
    except Exception:
        pass

    try:
        df_pl = pd.read_excel(polska_path)
    except Exception as e:
        print(f"[BŁĄD] Nie mogę wczytać Polska.xlsx: {polska_path}\n{e}")
        return 1

    col_area_pl = _find_col(df_pl.columns, ["metry", "powierzchnia", "Obszar", "obszar"])
    col_price_pl = _find_col(df_pl.columns, ["cena_za_metr", "cena za metr", "cena_za_m2", "cena_za_metr2", "cena za m2"])
    if not col_area_pl or not col_price_pl:
        print("[BŁĄD] Polska.xlsx nie zawiera wymaganych kolumn metrażu / ceny.")
        return 1

    # indeksy dla szybkich dopasowań (kanonizacja + mapy miejscowości)
    pl_index = build_polska_index(df_pl, col_area_pl, col_price_pl)

    is_excel = raport_path.suffix.lower() in [".xlsx", ".xlsm", ".xls"]
    try:
        if is_excel:
            sheet_to_read = _pick_report_sheet_name(raport_path, preferred="raport")
            df_raport = pd.read_excel(raport_path, sheet_name=sheet_to_read)
        else:
            df_raport = pd.read_csv(raport_path, sep=None, engine="python")
    except Exception as e:
        print(f"[BŁĄD] Nie mogę wczytać raportu: {raport_path}\n{e}")
        return 1

    local_ludnosc = _find_ludnosc_csv(baza_folder=baza_folder, raport_path=raport_path, polska_path=polska_path)
    api_cache = baza_folder / "population_cache.csv"

    print(f"[Automat] local ludnosc.csv -> {local_ludnosc if local_ludnosc else '(NIE ZNALEZIONO)'}")

    pop_resolver = PopulationResolver(local_csv=local_ludnosc, api_cache_csv=api_cache, use_api=True)

    print(f"[Automat] Start – liczba wierszy w raporcie: {len(df_raport.index)}")

    for idx in range(len(df_raport.index)):
        try:
            _process_row(
                df_raport=df_raport,
                idx=idx,
                pl=pl_index,
                margin_m2_default=margin_m2_default,
                margin_pct_default=margin_pct_default,
                pop_resolver=pop_resolver,
            )
        except Exception as e:
            print(f"[Automat] Błąd przy wierszu {idx+1}: {e}")

    try:
        if is_excel:
            save_report_sheet_only(raport_path, df_raport, sheet_name="raport")
        else:
            df_raport.to_csv(raport_path, index=False, encoding="utf-8-sig")
    except Exception as e:
        print(f"[BŁĄD] Nie udało się zapisać raportu: {raport_path}\n{e}")
        return 1

    print(f"[Automat] Zakończono – zapisano zmiany w pliku: {raport_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
